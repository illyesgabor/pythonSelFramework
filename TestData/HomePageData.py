import openpyxl


class HomePageData:

    test_homepage_data = [{"firstname": "Chance", "lastname": "Denis", "gender": "Male"},
                            {"firstname": "Colette", "lastname": "Smith", "gender": "Female"}]

    @staticmethod
    def get_test_data(test_case_name):
        dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Gabor\\PycharmProjects\\pythonSelFramework\\TestData\\DemoExcel.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict]
