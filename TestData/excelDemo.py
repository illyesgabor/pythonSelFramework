import openpyxl


book = openpyxl.load_workbook("C:\\Users\\Gabor\\PycharmProjects\\pythonSelFramework\\TestData\\DemoExcel.xlsx")
sheet = book.active
dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2, sheet.max_column+1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict)
# book.save("DemoExcel.xlsx")
